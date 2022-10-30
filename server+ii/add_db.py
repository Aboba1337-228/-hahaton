import openpyxl
wookbook = openpyxl.load_workbook("dataset.xlsx")
worksheet = wookbook.active

for i in range(0, worksheet.max_row):
    print("(")
    for col in worksheet.iter_cols(0, worksheet.max_column):
        print(f"'{col[i].value}', ", end="\t")
    print(" '0'),")